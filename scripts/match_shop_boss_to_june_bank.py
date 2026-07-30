import csv
import itertools
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path.home() / "Downloads" / "Shop Boss - Reconciliation Data.xlsx"
BANK = ROOT / "odoo_imports" / "bank_reconciliation" / "odoo_unreconciled_bank_statement_lines_laurel_bank_live.csv"
OUT = ROOT / "odoo_imports" / "bank_reconciliation"
MONTH = sys.argv[1] if len(sys.argv) > 1 else "2026-06"
MONTH_LABEL = MONTH.replace("-", "_")
SUMMARY = OUT / f"shop_boss_{MONTH_LABEL}_daily_totals.csv"
MATCHES = OUT / f"shop_boss_{MONTH_LABEL}_bank_match_candidates.csv"
FEE_RATE = Decimal("0.035")
SUBSET_MATCHES = OUT / f"shop_boss_{MONTH_LABEL}_bank_fee_adjusted_subset_candidates.csv"


def money(value):
    if value is None or value == "":
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def date_key(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if value:
        return str(value)[:10]
    return ""


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Shop Production Detail Report -"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: pos for pos, name in enumerate(headers)}

    rows = []
    by_date = defaultdict(lambda: {"count": 0, "total": Decimal("0.00"), "tax": Decimal("0.00"), "ro": 0, "parts": 0})
    by_date_type = defaultdict(lambda: {"count": 0, "total": Decimal("0.00")})
    by_date_amount = defaultdict(list)
    for values in ws.iter_rows(min_row=2, values_only=True):
        status_date = date_key(values[idx["Status Date"]])
        if not status_date.startswith(MONTH):
            continue
        row_type = values[idx["Column1"]] or ""
        total = money(values[idx["Total RO"]])
        tax = money(values[idx["Tax"]])
        record = {
            "Source Type": row_type,
            "Source Number": values[idx["RO#"]],
            "Date": status_date,
            "Customer": values[idx["Customer"]],
            "Total": total,
            "Tax": tax,
        }
        rows.append(record)
        by_date[status_date]["count"] += 1
        by_date[status_date]["total"] += total
        by_date[status_date]["tax"] += tax
        by_date[status_date]["ro"] += 1 if row_type == "Repair Orders" else 0
        by_date[status_date]["parts"] += 1 if row_type == "Part Sales" else 0
        by_date_type[(status_date, row_type)]["count"] += 1
        by_date_type[(status_date, row_type)]["total"] += total
        by_date_amount[(status_date, total)].append(record)

    bank_rows = []
    with BANK.open("r", newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if row.get("date", "").startswith(MONTH) and money(row.get("amount")) > 0:
                bank_rows.append(row)

    with SUMMARY.open("w", newline="", encoding="utf-8-sig") as f:
        fields = ["Date", "Records", "Repair Orders", "Part Sales", "Shop Boss Total", "Tax Included"]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for day in sorted(by_date):
            data = by_date[day]
            writer.writerow(
                {
                    "Date": day,
                    "Records": data["count"],
                    "Repair Orders": data["ro"],
                    "Part Sales": data["parts"],
                    "Shop Boss Total": float(data["total"]),
                    "Tax Included": float(data["tax"]),
                }
            )

    match_rows = []
    fee_adjusted_rows = []
    subset_rows = []
    for bank in bank_rows:
        bank_date = datetime.fromisoformat(bank["date"]).date()
        amount = money(bank["amount"])
        for offset in range(-3, 4):
            source_day = (bank_date + timedelta(days=offset)).isoformat()
            exact = by_date_amount.get((source_day, amount), [])
            for source in exact:
                match_rows.append(
                    {
                        "Match Type": "Exact record amount within +/- 3 days",
                        "Bank Line ID": bank["id"],
                        "Bank Date": bank["date"],
                        "Bank Ref": bank["payment_ref"],
                        "Bank Amount": float(amount),
                        "Source Date": source["Date"],
                        "Source Type": source["Source Type"],
                        "Source Number": source["Source Number"],
                        "Customer": source["Customer"],
                        "Source Total": float(source["Total"]),
                    }
                )
        for offset in range(-3, 4):
            source_day = (bank_date + timedelta(days=offset)).isoformat()
            daily = by_date.get(source_day)
            if daily and daily["total"] == amount:
                match_rows.append(
                    {
                        "Match Type": "Exact daily total within +/- 3 days",
                        "Bank Line ID": bank["id"],
                        "Bank Date": bank["date"],
                        "Bank Ref": bank["payment_ref"],
                        "Bank Amount": float(amount),
                        "Source Date": source_day,
                        "Source Type": "Daily total",
                        "Source Number": "",
                        "Customer": "",
                        "Source Total": float(daily["total"]),
                    }
                )
            if daily:
                gross = daily["total"]
                estimated_fee = (gross * FEE_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                estimated_net = (gross - estimated_fee).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                difference = (amount - estimated_net).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if abs(difference) <= Decimal("2.00"):
                    fee_adjusted_rows.append(
                        {
                            "Match Type": "Daily total less 3.5% fee within +/- 3 days",
                            "Bank Line ID": bank["id"],
                            "Bank Date": bank["date"],
                            "Bank Ref": bank["payment_ref"],
                            "Bank Amount": float(amount),
                            "Source Date": source_day,
                            "Source Type": "Daily total",
                            "Source Number": "",
                            "Customer": "",
                            "Source Gross": float(gross),
                            "Estimated Fee 3.5%": float(estimated_fee),
                            "Estimated Net": float(estimated_net),
                            "Difference": float(difference),
                        }
                    )

        # One-record net deposit match, useful when a card batch contains a single RO/sale.
        for offset in range(-3, 4):
            source_day = (bank_date + timedelta(days=offset)).isoformat()
            for source in rows:
                if source["Date"] != source_day:
                    continue
                gross = source["Total"]
                estimated_fee = (gross * FEE_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                estimated_net = (gross - estimated_fee).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                difference = (amount - estimated_net).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if abs(difference) <= Decimal("2.00"):
                    fee_adjusted_rows.append(
                        {
                            "Match Type": "Record total less 3.5% fee within +/- 3 days",
                            "Bank Line ID": bank["id"],
                            "Bank Date": bank["date"],
                            "Bank Ref": bank["payment_ref"],
                            "Bank Amount": float(amount),
                            "Source Date": source["Date"],
                            "Source Type": source["Source Type"],
                            "Source Number": source["Source Number"],
                            "Customer": source["Customer"],
                            "Source Gross": float(gross),
                            "Estimated Fee 3.5%": float(estimated_fee),
                            "Estimated Net": float(estimated_net),
                            "Difference": float(difference),
                        }
                    )

        # Batched merchant deposits: try combinations of same-day Shop Boss records.
        for offset in range(-3, 4):
            source_day = (bank_date + timedelta(days=offset)).isoformat()
            day_records = [row for row in rows if row["Date"] == source_day and row["Total"] > 0]
            if len(day_records) > 14:
                continue
            for size in range(2, len(day_records) + 1):
                for combo in itertools.combinations(day_records, size):
                    gross = sum((item["Total"] for item in combo), Decimal("0.00"))
                    estimated_fee = (gross * FEE_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    estimated_net = (gross - estimated_fee).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    difference = (amount - estimated_net).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    if abs(difference) <= Decimal("5.00"):
                        subset_rows.append(
                            {
                                "Bank Line ID": bank["id"],
                                "Bank Date": bank["date"],
                                "Bank Ref": bank["payment_ref"],
                                "Bank Amount": float(amount),
                                "Source Date": source_day,
                                "Record Count": len(combo),
                                "Source Gross": float(gross),
                                "Estimated Fee 3.5%": float(estimated_fee),
                                "Estimated Net": float(estimated_net),
                                "Difference": float(difference),
                                "Source Numbers": "; ".join(f"{item['Source Type']} {item['Source Number']}" for item in combo),
                                "Customers": "; ".join(str(item["Customer"]) for item in combo),
                            }
                        )

    with MATCHES.open("w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "Match Type",
            "Bank Line ID",
            "Bank Date",
            "Bank Ref",
            "Bank Amount",
            "Source Date",
            "Source Type",
            "Source Number",
            "Customer",
            "Source Total",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(match_rows)

    fee_matches = OUT / f"shop_boss_{MONTH_LABEL}_bank_fee_adjusted_candidates.csv"
    with fee_matches.open("w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "Match Type",
            "Bank Line ID",
            "Bank Date",
            "Bank Ref",
            "Bank Amount",
            "Source Date",
            "Source Type",
            "Source Number",
            "Customer",
            "Source Gross",
            "Estimated Fee 3.5%",
            "Estimated Net",
            "Difference",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(fee_adjusted_rows)

    subset_rows.sort(key=lambda row: (abs(Decimal(str(row["Difference"]))), row["Bank Date"], row["Bank Line ID"], row["Record Count"]))
    with SUBSET_MATCHES.open("w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "Bank Line ID",
            "Bank Date",
            "Bank Ref",
            "Bank Amount",
            "Source Date",
            "Record Count",
            "Source Gross",
            "Estimated Fee 3.5%",
            "Estimated Net",
            "Difference",
            "Source Numbers",
            "Customers",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(subset_rows)

    print(f"{MONTH} Shop Boss rows: {len(rows)}")
    print(f"{MONTH} positive bank lines: {len(bank_rows)}")
    print(f"Exact match candidates: {len(match_rows)}")
    print(f"Fee-adjusted candidates: {len(fee_adjusted_rows)}")
    print(f"Fee-adjusted subset candidates: {len(subset_rows)}")
    print(f"Summary: {SUMMARY}")
    print(f"Matches: {MATCHES}")
    print(f"Fee-adjusted matches: {fee_matches}")
    print(f"Fee-adjusted subset matches: {SUBSET_MATCHES}")
    print()
    print(f"{MONTH} Shop Boss daily totals:")
    for day in sorted(by_date):
        data = by_date[day]
        print(f"{day}: records={data['count']} total={data['total']} tax={data['tax']}")
    print()
    print("Exact matches:")
    for row in match_rows[:50]:
        print(row)
    print()
    print("Fee-adjusted matches:")
    for row in fee_adjusted_rows[:80]:
        print(row)
    print()
    print("Best fee-adjusted subset matches:")
    for row in subset_rows[:80]:
        print(row)


if __name__ == "__main__":
    main()
