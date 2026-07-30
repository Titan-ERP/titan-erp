from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path.home() / "Downloads" / "Shop Boss - Reconciliation Data.xlsx"


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    print(f"Workbook: {SOURCE}")
    print(f"Sheets: {wb.sheetnames}")
    for ws in wb.worksheets:
        print()
        print(f"SHEET: {ws.title} rows={ws.max_row} cols={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            print(row)


if __name__ == "__main__":
    main()
