import csv
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path.home() / "Downloads" / "Shop Boss - Reconciliation Data.xlsx"
OUT = ROOT / "odoo_imports" / "bank_reconciliation"
EXACT = OUT / "shop_boss_june_bank_match_candidates.csv"
FEE_SUBSET = OUT / "shop_boss_june_bank_fee_adjusted_subset_candidates.csv"
PLAN = OUT / "shop_boss_june_deposit_reconcile_review_plan.csv"
FEE_RATE = Decimal("0.035")


def money(value):
    if value is None or value == "":
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def load_shop_boss_rows():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Shop Production Detail Report -"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: pos for pos, name in enumerate(headers)}
    rows = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        source_type = values[idx["Column1"]]
        source_number = values[idx["RO#"]]
        if not source_type or source_number is None:
            continue
        key = f"{source_type} {source_number}"
        rows[key] = {
            "Source Key": key,
            "Source Type": source_type,
            "Source Number": source_number,
            "Customer": values[idx["Customer"]],
            "Taxable Labor": money(values[idx["Taxable Labor"]]),
            "Non-Tax Labor": money(values[idx["Non-Tax Labor"]]),
            "Taxable Parts": money(values[idx["Taxable Parts"]]),
            "Non-Tax Parts": money(values[idx["Non-Tax Parts"]]),
            "Sublet": money(values[idx["Sublet"]]),
            "Fees": money(values[idx["Fees"]]),
            "Tax": money(values[idx["Tax"]]),
            "Discount": money(values[idx["Discount"]]),
            "Total RO": money(values[idx["Total RO"]]),
            "Parts Cost": money(values[idx["Parts Cost"]]),
        }
    return rows


def allocation(records):
    labor = sum((row["Taxable Labor"] + row["Non-Tax Labor"] for row in records), Decimal("0.00"))
    parts = sum((row["Taxable Parts"] + row["Non-Tax Parts"] for row in records), Decimal("0.00"))
    sublet = sum((row["Sublet"] for row in records), Decimal("0.00"))
    fees = sum((row["Fees"] for row in records), Decimal("0.00"))
    tax = sum((row["Tax"] for row in records), Decimal("0.00"))
    discount = sum((row["Discount"] for row in records), Decimal("0.00"))
    gross = sum((row["Total RO"] for row in records), Decimal("0.00"))
    return {
        "Service Revenue": labor,
        "Parts Revenue": parts,
        "Other Revenue/Fees": sublet + fees,
        "Sales Tax Payable": tax,
        "Discount": discount,
        "Gross": gross,
    }


def parse_keys(value):
    return [part.strip() for part in str(value or "").split(";") if part.strip()]


def candidate_rows():
    rows = []
    if EXACT.exists():
        with EXACT.open("r", newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                rows.append(
                    {
                        "Match Source": "Exact",
                        "Bank Line ID": row["Bank Line ID"],
                        "Bank Date": row["Bank Date"],
                        "Bank Ref": row["Bank Ref"],
                        "Bank Amount": money(row["Bank Amount"]),
                        "Source Date": row["Source Date"],
                        "Source Keys": [f"{row['Source Type']} {row['Source Number']}"],
                        "Difference": Decimal("0.00"),
                    }
                )
    if FEE_SUBSET.exists():
        with FEE_SUBSET.open("r", newline="", encoding="utf-8-sig") as f:
            seen_bank = set()
            for row in csv.DictReader(f):
                bank_id = row["Bank Line ID"]
                diff = money(row["Difference"])
                if abs(diff) > Decimal("1.50"):
                    continue
                # Keep only the best low-difference candidate per bank line for review.
                if bank_id in seen_bank:
                    continue
                seen_bank.add(bank_id)
                rows.append(
                    {
                        "Match Source": "Fee Adjusted Subset",
                        "Bank Line ID": bank_id,
                        "Bank Date": row["Bank Date"],
                        "Bank Ref": row["Bank Ref"],
                        "Bank Amount": money(row["Bank Amount"]),
                        "Source Date": row["Source Date"],
                        "Source Keys": parse_keys(row["Source Numbers"]),
                        "Difference": diff,
                    }
                )
    return rows


def main():
    shop_rows = load_shop_boss_rows()
    review_rows = []
    for candidate in candidate_rows():
        records = [shop_rows[key] for key in candidate["Source Keys"] if key in shop_rows]
        if not records:
            continue
        alloc = allocation(records)
        gross = alloc["Gross"]
        estimated_fee = Decimal("0.00") if candidate["Match Source"] == "Exact" else (gross * FEE_RATE).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        estimated_net = (gross - estimated_fee).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        review_rows.append(
            {
                "Review Status": "Ready for review",
                "Match Source": candidate["Match Source"],
                "Bank Line ID": candidate["Bank Line ID"],
                "Bank Date": candidate["Bank Date"],
                "Bank Ref": candidate["Bank Ref"],
                "Bank Amount": float(candidate["Bank Amount"]),
                "Source Date": candidate["Source Date"],
                "Source Records": "; ".join(candidate["Source Keys"]),
                "Customers": "; ".join(str(row["Customer"]) for row in records),
                "Gross Shop Boss": float(gross),
                "Estimated Merchant Fee": float(estimated_fee),
                "Estimated Net Deposit": float(estimated_net),
                "Bank vs Estimated Net": float((candidate["Bank Amount"] - estimated_net).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
                "Credit Parts Revenue": float(alloc["Parts Revenue"]),
                "Credit Service Revenue": float(alloc["Service Revenue"]),
                "Credit Other Revenue/Fees": float(alloc["Other Revenue/Fees"]),
                "Credit Sales Tax Payable": float(alloc["Sales Tax Payable"]),
                "Debit Bank Merchant Fees": float(estimated_fee),
                "Notes": "Review before applying. Net deposits need fee split; exact deposits may be cash/check or no-fee card settlement.",
            }
        )

    with PLAN.open("w", newline="", encoding="utf-8-sig") as f:
        fields = [
            "Review Status",
            "Match Source",
            "Bank Line ID",
            "Bank Date",
            "Bank Ref",
            "Bank Amount",
            "Source Date",
            "Source Records",
            "Customers",
            "Gross Shop Boss",
            "Estimated Merchant Fee",
            "Estimated Net Deposit",
            "Bank vs Estimated Net",
            "Credit Parts Revenue",
            "Credit Service Revenue",
            "Credit Other Revenue/Fees",
            "Credit Sales Tax Payable",
            "Debit Bank Merchant Fees",
            "Notes",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(review_rows)

    print(f"Review rows: {len(review_rows)}")
    print(f"Plan: {PLAN}")
    for row in review_rows:
        print(row)


if __name__ == "__main__":
    main()
