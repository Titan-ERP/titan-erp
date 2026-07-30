import csv
import os
import re
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
IN_LINES = ROOT / "odoo_imports" / "bank_reconciliation" / "odoo_unreconciled_bank_statement_lines_laurel_bank_live.csv"
OUT = ROOT / "odoo_imports" / "bank_reconciliation"
WORKBOOK = OUT / "bank_reconciliation_review_laurel_bank_live.xlsx"
GROUP_CSV = OUT / "bank_reconciliation_unreconciled_groups.csv"


ACCOUNT_RULES = [
    (r"BANKCARD.*MTOT DEP|DEPOSIT", "Undeposited Funds / Card Clearing", "Likely customer/card deposit - match to POS/Sales batch before coding."),
    (r"MONTHLY DEBIT CARD FEE|SERVICE CHARGE|ANALYSIS CHARGE|BANK FEE", "Bank Fees", "Bank fee."),
    (r"IRS|MSDEPTOFREVENUE|TAXPAYMENT", "Taxes Payable", "Tax payment - verify payroll/sales tax liability account."),
    (r"INTUIT PAYROLL|QUICKBOOKS", "Payroll Expense / Payroll Clearing", "Payroll processor - verify payroll journal/liability setup."),
    (r"TELEPHONE TRF TO LN|ATS - CHECKING TO LN", "Loan Payable", "Loan payment/transfer - verify principal vs interest split."),
    (r"AMAZON|OFFICE DEPOT|STAPLES", "Office Supplies", "Office/admin purchase."),
    (r"MURPHY USA|TEXACO|EXXON|SHELL|CHEVRON|FUEL", "Fuel Expense", "Fuel purchase."),
    (r"UPS|FEDEX|USPS", "Shipping Expense", "Shipping/freight."),
    (r"SPAREX|TRACTORPARTS|AICPARTS|TRACTO PARTS|B&M TRACTOR|PAINT VALLEY|CROSS CREEK TRACTOR|FINNEY EQUIPMENT|PUCKETT MACHINERY|COWIN EQUIPMENT|GRAINGER|HB SEALING|HANDR AGRI", "Parts COGS", "Parts/vendor purchase - verify if inventory bill exists."),
    (r"PANDA EXPRESS|RESTAURANT|MCDONALD|CHICK-FIL-A|FOOD", "Meals Expense", "Meals - verify business purpose."),
]


def money(value):
    return Decimal(str(value or "0")).quantize(Decimal("0.01"))


def merchant_key(ref):
    text = str(ref or "").upper()
    text = re.sub(r"\*+\d+", " ", text)
    text = re.sub(r"\b\d{2}/\d{2}\b.*$", " ", text)
    text = re.sub(r"\b\d{1,2}:\d{2}\b", " ", text)
    text = re.sub(r"\b(POS PURCHASE|NON PIN|PIN|DEBIT|CARD|PURCHASE|CHECKCARD)\b", " ", text)
    text = re.sub(r"[^A-Z0-9&/# -]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return " ".join(text.split()[:6]) or "Unclear"


def suggestion(ref):
    text = str(ref or "").upper()
    for pattern, account, note in ACCOUNT_RULES:
        if re.search(pattern, text):
            return account, note
    if "CHECK" in text:
        return "Needs Review", "Check payment - needs check number/payee detail."
    return "Needs Review", "No safe rule. Review manually."


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def main():
    rows = []
    with IN_LINES.open("r", newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            account, note = suggestion(row.get("payment_ref", ""))
            amount = money(row.get("amount"))
            row["amount_number"] = float(amount)
            row["merchant_key"] = merchant_key(row.get("payment_ref", ""))
            row["suggested_account"] = account
            row["suggestion_note"] = note
            rows.append(row)

    group_map = defaultdict(list)
    for row in rows:
        group_map[(row["merchant_key"], row["suggested_account"], row["suggestion_note"])].append(row)

    groups = []
    for (key, account, note), group_rows in group_map.items():
        groups.append(
            {
                "merchant_key": key,
                "count": len(group_rows),
                "total_amount": float(sum(money(row["amount"]) for row in group_rows)),
                "suggested_account": account,
                "suggestion_note": note,
                "example_ref": group_rows[0].get("payment_ref", ""),
            }
        )
    groups.sort(key=lambda row: (-row["count"], row["merchant_key"]))

    with GROUP_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["merchant_key", "count", "total_amount", "suggested_account", "suggestion_note", "example_ref"],
        )
        writer.writeheader()
        writer.writerows(groups)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Groups"
    group_headers = ["merchant_key", "count", "total_amount", "suggested_account", "suggestion_note", "example_ref"]
    ws.append(group_headers)
    for group in groups:
        ws.append([group[h] for h in group_headers])

    detail = wb.create_sheet("Unreconciled Lines")
    headers = [
        "id",
        "date",
        "display_name",
        "payment_ref",
        "partner_id",
        "amount",
        "journal_id",
        "merchant_key",
        "suggested_account",
        "suggestion_note",
    ]
    detail.append(headers)
    for row in rows:
        detail.append([row.get(h, "") if h != "amount" else row.get("amount_number", "") for h in headers])

    summary = wb.create_sheet("Summary")
    total = sum(money(row["amount"]) for row in rows)
    by_account = Counter(row["suggested_account"] for row in rows)
    summary.append(["Metric", "Value"])
    summary.append(["Unreconciled lines", len(rows)])
    summary.append(["Net unreconciled amount", float(total)])
    summary.append(["Groups", len(groups)])
    summary.append([])
    summary.append(["Suggested Account", "Lines"])
    for account, count in by_account.most_common():
        summary.append([account, count])

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F2D59")
        sheet.freeze_panes = "A2"
        autosize(sheet)

    wb.save(WORKBOOK)
    print(f"Workbook: {WORKBOOK}")
    print(f"Groups CSV: {GROUP_CSV}")
    print(f"Unreconciled lines: {len(rows)}")
    print(f"Groups: {len(groups)}")
    for account, count in by_account.most_common():
        print(f"{account}: {count}")


if __name__ == "__main__":
    main()
