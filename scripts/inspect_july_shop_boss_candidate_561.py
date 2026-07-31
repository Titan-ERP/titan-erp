from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path.home() / "Downloads" / "Shop Boss - Reconciliation Data.xlsx"
KEYS = {
    ("Repair Orders", 1108),
    ("Repair Orders", 1111),
    ("Part Sales", 394),
    ("Part Sales", 397),
    ("Part Sales", 407),
}


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Shop Production Detail Report -"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {value: pos for pos, value in enumerate(headers)}
    fields = ["Taxable Labor", "Non-Tax Labor", "Taxable Parts", "Non-Tax Parts", "Sublet", "Fees", "Tax", "Total RO"]
    sums = {field: Decimal("0.00") for field in fields}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = (row[idx["Column1"]], row[idx["RO#"]])
        if key not in KEYS:
            continue
        print(key, row[idx["Status Date"]], row[idx["Customer"]], row[idx["Total RO"]], "tax", row[idx["Tax"]])
        for field in fields:
            sums[field] += Decimal(str(row[idx[field]] or 0))
    print("SUMS")
    for field, value in sums.items():
        print(field, value)
    print("Parts Revenue", sums["Taxable Parts"] + sums["Non-Tax Parts"])
    print("Service Revenue", sums["Taxable Labor"] + sums["Non-Tax Labor"])
    print("Other Revenue/Fees", sums["Sublet"] + sums["Fees"])
    print("Gross", sums["Total RO"])
    print("Actual Merchant Fee", sums["Total RO"] - Decimal("1441.82"))


if __name__ == "__main__":
    main()
